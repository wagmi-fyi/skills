#!/usr/bin/env python3
"""
Export to Excel — Format Adapter
Transforms universal JSON from extract_review_items.py into a formatted .xlsx workbook.
Pure format transform — no database operations.

Usage:
    # Piped from extractor
    python scripts/extract_review_items.py --type=client_questions --start_date="2024-01-01" --end_date="2024-01-31" --period_label="2024-01" | \
      python adapters/export_to_excel.py --output="/tmp/client_questions.xlsx"

    # From saved JSON file
    python adapters/export_to_excel.py --input="/tmp/review_items.json" --output="/tmp/review.xlsx"
"""

import argparse
import json
import os
import sys
import traceback

# Add scripts/_shared to Python path for config_loader
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts', '_shared'))
import config_loader

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl package required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Column Layout Definitions
# ---------------------------------------------------------------------------
# Each column: (header, source_field_or_None, is_fill_in)

COLUMN_LAYOUTS = {
    "client_questions": [
        ("Date",           "date",               False),
        ("Amount",         "amount_display",      False),
        ("Reference",      "reference",           False),
        ("Account",        "source",              False),
        ("Question",       "suggested_question",  False),
        ("Context",        "context",             False),
        ("Client Answer",  None,                  True),
        ("Account Code",   None,                  True),
    ],
    "judgment_calls": [
        ("Date",             "date",                     False),
        ("Amount",           "amount_display",            False),
        ("Reference",        "reference",                 False),
        ("Account",          "source",                    False),
        ("AI Category",      "categorized_account_name",  False),
        ("AI Account Code",  "categorized_account_code",  False),
        ("Confidence",       "confidence_score",           False),
        ("Contact",          "contact",                    False),
        ("Approved?",        None,                         True),
        ("Correct Account",  None,                         True),
    ],
    "review_dashboard": [
        ("Check",          "check_name",    False),
        ("Status",         "status",        False),
        ("Items Flagged",  "flagged_count", False),
        ("Notes",          "notes",         False),
    ],
    "transaction_register": [
        ("Date",           "date",          False),
        ("Account Code",   "account_code",  False),
        ("Account Name",   "account_name",  False),
        ("Reference",      "reference",     False),
        ("Contact",        "contact",       False),
        ("Memo",           "memo",          False),
        ("Debit",          "debit",         False),
        ("Credit",         "credit",        False),
        ("Balance",        "balance",       False),
        ("Confidence",     "confidence_score", False),  # optional field; judgment-zone rows highlight yellow
    ],
    "subledger_gl_tie": [
        ("Description",      "description",      False),
        ("Subledger Total",  "subledger_total",  False),
        ("GL Balance",       "gl_balance",       False),
        ("Difference",       "difference",       False),
        ("Status",           "status",           False),
    ],
    "pop_variance": [
        ("Account Code",      "account_code",      False),
        ("Account Name",      "account_name",      False),
        ("Type",              "account_type",      False),
        ("Prior Period",      "prior_period",      False),
        ("Current Period",    "current_period",    False),
        ("% Rev (Current)",   "pct_rev_current",   False),
        ("% Rev (Prior)",     "pct_rev_prior",     False),
        ("Shift",             "shift",             False),
        ("Severity",          "severity",          False),
    ],
}

SHEET_NAMES = {
    "client_questions": "Client Questions",
    "judgment_calls": "Judgment Calls",
    "review_dashboard": "Dashboard",
    "transaction_register": "Transaction Register",
    "subledger_gl_tie": "Subledger-GL Tie",
    "pop_variance": "PoP Variance",
}

# Formatting constants
FILL_IN_FILL = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
CONFIDENCE_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
CONFIDENCE_GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
HEADER_FONT = Font(bold=True)
SUMMARY_FONT = Font(bold=True, size=14)
RIGHT_ALIGN = Alignment(horizontal="right")
MAX_COL_WIDTH = 50


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_arguments():
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(
        description="Export universal JSON from extract_review_items.py to formatted Excel workbook"
    )
    parser.add_argument("--input", default=None, help="Path to JSON file (reads from stdin if omitted)")
    parser.add_argument("--output", required=True, help="Output path for Excel file (.xlsx)")
    parser.add_argument("--template", default=None, help="Path to Excel template for client branding (future)")
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Input Loading
# ---------------------------------------------------------------------------

def load_input(args):
    """Load and validate JSON input from file or stdin."""
    if args.input:
        with open(args.input, "r") as f:
            raw = f.read()
    else:
        raw = sys.stdin.read()

    if not raw.strip():
        raise ValueError("Empty input — provide JSON via --input or stdin")

    data = json.loads(raw)

    # Validate required fields
    if "type" not in data:
        raise ValueError("Invalid input JSON: missing 'type' field")

    # review_package is a multi-sheet envelope — validate differently
    if data["type"] == "review_package":
        if "sheets" not in data or not isinstance(data["sheets"], list):
            raise ValueError("review_package requires a 'sheets' array")
        for i, sheet in enumerate(data["sheets"]):
            if "type" not in sheet or "items" not in sheet:
                raise ValueError(f"Sheet {i} missing 'type' or 'items'")
            if sheet["type"] not in COLUMN_LAYOUTS:
                raise ValueError(
                    f"Sheet {i} has unsupported type '{sheet['type']}'. "
                    f"Must be one of: {', '.join(COLUMN_LAYOUTS.keys())}"
                )
        return data

    if "items" not in data:
        raise ValueError("Invalid input JSON: missing 'items' field")
    if data["type"] not in COLUMN_LAYOUTS:
        raise ValueError(
            f"Unsupported type '{data['type']}'. Must be one of: {', '.join(COLUMN_LAYOUTS.keys())}"
        )

    return data


# ---------------------------------------------------------------------------
# Summary Header
# ---------------------------------------------------------------------------

def write_summary_header(ws, data, num_columns):
    """Write merged summary row at top of sheet."""
    data_type = data["type"]
    type_display = SHEET_NAMES.get(data_type, data_type)
    period = data.get("period_label", "All")
    count = data.get("count", len(data["items"]))

    # Compute date range from items
    items = data["items"]
    dates = [item["date"] for item in items if item.get("date")]
    if dates:
        date_range = f"{min(dates)} to {max(dates)}"
    else:
        date_range = "N/A"

    summary = f"{type_display} | Period: {period} | Count: {count} | Date range: {date_range}"

    # Merge across all columns in row 1
    if num_columns > 1:
        end_col = get_column_letter(num_columns)
        ws.merge_cells(f"A1:{end_col}1")

    cell = ws.cell(row=1, column=1, value=summary)
    cell.font = SUMMARY_FONT


# ---------------------------------------------------------------------------
# Data Rows + Formatting
# ---------------------------------------------------------------------------

def write_data_rows(ws, items, columns, data_type):
    """Write column headers and data rows with formatting."""
    header_row = 2
    data_start_row = 3

    # -- Column headers (row 2) --
    for col_idx, (header, _source, _fill_in) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT

    # -- Freeze pane: keep rows 1-2 visible --
    ws.freeze_panes = "A3"

    # -- Data rows --
    for row_offset, item in enumerate(items):
        row_num = data_start_row + row_offset
        for col_idx, (header, source_field, is_fill_in) in enumerate(columns, start=1):
            if is_fill_in:
                # Empty fill-in column with highlight
                cell = ws.cell(row=row_num, column=col_idx, value="")
                cell.fill = FILL_IN_FILL
            else:
                # confidence_score is optional (older envelope producers omit it)
                if source_field == "confidence_score":
                    value = item.get(source_field)
                else:
                    value = item[source_field]
                cell = ws.cell(row=row_num, column=col_idx, value=value)

            # Right-align Amount column (strings won't auto-right-align)
            if header == "Amount":
                cell.alignment = RIGHT_ALIGN

            # Conditional formatting for Confidence column (judgment_calls)
            # Yellow = lower half of judgment range, Green = upper half
            if header == "Confidence" and not is_fill_in and data_type == "judgment_calls":
                score = item.get("confidence_score")
                if score is not None:
                    coding_cfg = config_loader.get_coding_config()
                    jc_min = coding_cfg['min_confidence_to_categorize']
                    jc_max = coding_cfg['min_confidence_to_auto_approve'] - 1
                    midpoint = jc_min + (jc_max - jc_min) // 2
                    if jc_min <= score <= midpoint:
                        cell.fill = CONFIDENCE_YELLOW
                    elif midpoint < score <= jc_max:
                        cell.fill = CONFIDENCE_GREEN

        # Row-level judgment-zone coloring for transaction_register:
        # rows whose confidence falls in the judgment range highlight yellow
        # (same treatment as pop_variance) so a reviewer can spot/filter them inline.
        if data_type == "transaction_register":
            score = item.get("confidence_score")
            if score is not None:
                coding_cfg = config_loader.get_coding_config()
                jc_min = coding_cfg['min_confidence_to_categorize']
                jc_max = coding_cfg['min_confidence_to_auto_approve'] - 1
                if jc_min <= score <= jc_max:
                    for col_idx in range(1, len(columns) + 1):
                        ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                            start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Row-level severity coloring for pop_variance
        if data_type == "pop_variance":
            severity = item.get("severity", "")
            if severity == "red":
                for col_idx in range(1, len(columns) + 1):
                    ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                        start_color="FF4444", end_color="FF4444", fill_type="solid")
                    ws.cell(row=row_num, column=col_idx).font = Font(color="FFFFFF", bold=True)
            elif severity == "yellow":
                for col_idx in range(1, len(columns) + 1):
                    ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                        start_color="FFC000", end_color="FFC000", fill_type="solid")

    # -- Auto-width (skip row 1 merged summary) --
    for col_idx, (header, _source, _fill_in) in enumerate(columns, start=1):
        max_length = len(str(header))
        for row_num in range(data_start_row, data_start_row + len(items)):
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    try:
        args = parse_arguments()

        # Warn on --template usage
        if args.template:
            print(
                "Warning: --template is not yet implemented. Template will be ignored.",
                file=sys.stderr,
            )

        # Load and validate input
        data = load_input(args)
        data_type = data["type"]

        # Create workbook
        wb = Workbook()

        if data_type == "review_package":
            # Multi-sheet review package
            sheet_names_out = []
            total_rows = 0
            for i, sheet_def in enumerate(data["sheets"]):
                s_type = sheet_def["type"]
                s_items = sheet_def["items"]
                s_columns = COLUMN_LAYOUTS[s_type]
                s_name = SHEET_NAMES[s_type]

                if i == 0:
                    ws = wb.active
                    ws.title = s_name
                else:
                    ws = wb.create_sheet(title=s_name)

                # Build a sheet-level data dict for summary header
                sheet_data = {
                    "type": s_type,
                    "items": s_items,
                    "period_label": data.get("period_label", ""),
                }
                write_summary_header(ws, sheet_data, len(s_columns))
                write_data_rows(ws, s_items, s_columns, s_type)
                sheet_names_out.append(s_name)
                total_rows += len(s_items)

            wb.save(args.output)

            result = {
                "success": True,
                "output_file": args.output,
                "type": "review_package",
                "rows": total_rows,
                "sheets": sheet_names_out,
            }
        else:
            # Single-sheet export (existing behavior)
            columns = COLUMN_LAYOUTS[data_type]
            items = data["items"]
            sheet_name = SHEET_NAMES[data_type]

            ws = wb.active
            ws.title = sheet_name

            write_summary_header(ws, data, len(columns))
            write_data_rows(ws, items, columns, data_type)

            wb.save(args.output)

            result = {
                "success": True,
                "output_file": args.output,
                "type": data_type,
                "rows": len(items),
                "sheets": [sheet_name],
            }
        print(json.dumps(result, indent=2))
        sys.exit(0)

    except Exception as e:
        traceback.print_exc(file=sys.stderr)
        print(json.dumps({"success": False, "error": str(e)}, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()
